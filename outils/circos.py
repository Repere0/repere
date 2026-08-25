#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Commune -> circonscription(s) legislative(s).

CE QUE CETTE TABLE PERMET : dire a quelqu'un dans quelle circonscription il vote.
Sans elle, l'application affiche les SIX deputes des Pyrenees-Atlantiques a un
habitant d'Ustaritz, sans pouvoir dire lequel est le sien. C'est le dernier
maillon de la chaine « du conseil municipal a l'Assemblee », et il manquait :
la table CIRCOS ecrite a la main dans l'application couvrait UNE commune.

CE QU'ELLE NE PERMET PAS ENCORE : nommer LE depute. Le rattachement d'un depute
a un numero de circonscription n'est pas dans le Repertoire national des elus —
mesure faite le 25/08/2026 : ses 577 lignes « Depute » portent un departement,
un nom, une fonction, et rien d'autre. Ce lien vit dans le referentiel des
acteurs de l'Assemblee (AMO30), que la chaine quotidienne telecharge deja.
Tant que la jointure n'est pas faite, l'application dit la circonscription et
ne nomme personne : c'est un fait de moins, pas un fait faux.

CE QU'ELLE NE PERMETTRA JAMAIS : designer LE depute d'une commune a cheval sur
plusieurs circonscriptions. Determiner laquelle demanderait l'adresse de la
personne, que Repere ne demandera jamais. Mesure : 34 508 communes sur 34 626
n'ont qu'une seule circonscription, 118 en ont plusieurs, dont Paris (18),
Marseille (7) et Toulouse (5). Pour celles-la, on les affiche TOUTES et on dit
que la commune est partagee.

SOURCE RETENUE, ET POURQUOI. Deux sources etaient collectees pour etre
comparees. C'est le XLSX du ministere de l'Interieur qui est retenu :
provenance directe, licence ouverte, et couverture mesuree de 99,97 % des
34 637 communes que l'application connait. Le CSV par bureau de vote, republie
par un tiers a partir des resultats de 2022, n'apporte pas assez pour justifier
une provenance indirecte sur une donnee civique. Il reste accepte en entree.

CE QUE LE FICHIER DU MINISTERE NE DIT PAS SIMPLEMENT. Il n'ecrit pas de code
INSEE : il ecrit un code de departement dans la convention du ministere (« ZA »
pour la Guadeloupe, « ZM » pour Mayotte) et un code de commune sur trois
chiffres, relatif au territoire. La reconstruction est donc explicite,
territoire par territoire, et VERIFIEE : chaque code reconstruit doit exister
dans la liste des communes que l'application connait, et les huit collectivites
d'outre-mer sont couvertes a 100 % — c'est ce qui prouve que les regles sont
justes plutot que plausibles.

LES ONZE COMMUNES ABSENTES. Le decoupage date de 2010 et le fichier de 2017.
Onze communes nouvelles creees depuis n'y figurent pas. Elles sont nommees dans
la sortie, pour que l'application dise « nous ne savons pas encore » au lieu de
se taire.

Usage :  python3 outils/circos.py <source.xlsx|source.csv> outils/circos.json [app.html]
"""
import sys, io, os, csv, json, collections

SRC = sys.argv[1] if len(sys.argv) > 1 else "data/circos_ministere.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "outils/circos.json"
APP = sys.argv[3] if len(sys.argv) > 3 else None

if not os.path.exists(SRC):
    sys.exit("source absente : %s — la collecte ne l'a pas telechargee." % SRC)


def communes_de_reference(chemin_app):
    """Les communes que l'application connait, relues dans le bloc RNE : code ->
       libelle. Sert de juge (un code reconstruit qui n'y figure pas est un code
       faux) ET de source des noms — les onze communes absentes du decoupage n'ont
       evidemment pas de nom dans le decoupage, il faut le prendre ici. Sans ca,
       la sortie annoncait des communes nommees et n'en nommait aucune."""
    if not chemin_app or not os.path.exists(chemin_app):
        return None
    import re
    with io.open(chemin_app, encoding="utf-8") as f:
        for ligne in f:
            if ligne.startswith("window.REPERE_RNE = "):
                m = re.search(r"=\s*(\{.*\});\s*$", ligne.strip())
                if m:
                    return json.loads(m.group(1))["cl"]
    return None


# Conventions du ministere -> code INSEE. Chacune est VERIFIEE plus bas contre la
# liste de reference ; aucune n'est supposee juste parce qu'elle est vraisemblable.
def insee_depuis_ministere(dpt, com):
    d = str(dpt).strip()
    c = int(com)
    if d in ("2A", "2B"):
        return d + "%03d" % c
    if d.isdigit():
        return "%02d%03d" % (int(d), c)
    if d in ("ZA", "ZB", "ZC", "ZD", "ZS"):   # Antilles, Guyane, Reunion, St-Pierre
        return "97%03d" % c                    # le code porte deja le chiffre du territoire
    if d == "ZM":                              # Mayotte : numerotee 5xx dans ce fichier
        return "976%02d" % (c - 500)
    if d == "ZN":                              # Nouvelle-Caledonie : 8xx
        return "988%02d" % (c - 800)
    if d == "ZP":                              # Polynesie francaise
        return "987%02d" % c
    # ZW (Wallis-et-Futuna), ZX (Saint-Martin / Saint-Barthelemy) et ZZ (Francais
    # etablis hors de France) ne sont pas des communes du Code officiel geographique.
    # ZZ en particulier porte les onze circonscriptions de l'etranger : les ranger
    # sous une commune serait faux.
    return None


def lire_xlsx(chemin):
    import openpyxl
    wb = openpyxl.load_workbook(chemin, read_only=True)
    ws = wb[wb.sheetnames[0]]
    lignes = ws.iter_rows(min_row=1, values_only=True)
    tete = [str(x or "").strip() for x in next(lignes)]
    besoin = ["CODE DPT", "CODE COMMUNE", "NOM COMMUNE", "CODE CIRC LEGISLATIVE"]
    manquantes = [c for c in besoin if c not in tete]
    assert not manquantes, "colonnes absentes du XLSX : %s (vues : %s)" % (manquantes, tete)
    i = {c: tete.index(c) for c in besoin}
    for r in lignes:
        if r[i["CODE DPT"]] is None:
            continue
        yield (insee_depuis_ministere(r[i["CODE DPT"]], r[i["CODE COMMUNE"]]),
               str(r[i["NOM COMMUNE"]] or "").strip(),
               int(r[i["CODE CIRC LEGISLATIVE"]]))


def lire_csv(chemin):
    """Le CSV par bureau de vote, garde en entree acceptee. Le delimiteur se devine
       sur l'en-tete : l'ecrire en dur etait exactement le defaut que la methode
       « lire avant d'ecrire » devait empecher."""
    tete = io.open(chemin, encoding="utf-8", errors="replace").readline()
    delim = max([";", ",", "\t", "|"], key=lambda d: tete.count(d))
    with io.open(chemin, encoding="utf-8", errors="replace", newline="") as f:
        lect = csv.DictReader(f, delimiter=delim)
        besoin = ["codeCommune", "codeCirconscription"]
        manquantes = [c for c in besoin if c not in (lect.fieldnames or [])]
        assert not manquantes, ("colonnes absentes : %s (delimiteur devine : %r, vues : %s)"
                                % (manquantes, delim, lect.fieldnames))
        for l in lect:
            code = (l["codeCommune"] or "").strip()
            circ = (l["codeCirconscription"] or "").strip()
            if not code or not circ:
                continue
            try:
                yield code, (l.get("nomCommune") or "").strip(), int(circ)
            except ValueError:
                continue


LIBELLES = communes_de_reference(APP)
REF = set(LIBELLES) if LIBELLES is not None else None
source_est_xlsx = SRC.lower().endswith(".xlsx")
lignes = lire_xlsx(SRC) if source_est_xlsx else lire_csv(SRC)

par_commune = collections.defaultdict(set)
noms = {}
n_lues = 0
ecartees = collections.Counter()
inconnues = collections.Counter()

for code, nom, circ in lignes:
    n_lues += 1
    if code is None:
        ecartees["hors Code officiel geographique"] += 1
        continue
    if REF is not None and code not in REF:
        # Commune fusionnee depuis 2017, ou code faux. La liste de reference tranche.
        inconnues[code[:2]] += 1
        continue
    par_commune[code].add(circ)
    noms.setdefault(code, nom)

assert n_lues > 30000, "seulement %d lignes lues : fichier tronque ?" % n_lues
assert len(par_commune) > 30000, "seulement %d communes rattachees" % len(par_commune)

# --------------------------------------------------------- controles independants
if REF is not None:
    couverture = len(par_commune) / float(len(REF))
    assert couverture > 0.99, "couverture de %.2f %% seulement" % (100 * couverture)
    # Les collectivites d'outre-mer sont le seul endroit ou la reconstruction du code
    # INSEE est une REGLE et non une concatenation. Si une seule regle etait fausse,
    # son territoire tomberait a zero : c'est ce que ce controle mesure.
    for prefixe, lib in (("971", "Guadeloupe"), ("972", "Martinique"), ("973", "Guyane"),
                         ("974", "La Reunion"), ("975", "Saint-Pierre-et-Miquelon"),
                         ("976", "Mayotte"), ("987", "Polynesie francaise"),
                         ("988", "Nouvelle-Caledonie")):
        attendues = [c for c in REF if c.startswith(prefixe)]
        trouvees = [c for c in attendues if c in par_commune]
        assert not attendues or trouvees, \
            "%s : aucune de ses %d communes rattachee — la regle de code INSEE est fausse" \
            % (lib, len(attendues))
    absentes = sorted(REF - set(par_commune))
else:
    absentes = []

# ------------------------------------------------------------------------- sortie
# Forme compacte : un entier quand la commune n'a qu'une circonscription, une liste
# triee sinon. 34 508 communes sur 34 626 sont dans le premier cas.
table = {}
for code, circs in par_commune.items():
    l = sorted(circs)
    table[code] = l[0] if len(l) == 1 else l

paquet = {
    "v": 1,
    "source": ("Ministere de l'Interieur — communes et cantons par circonscription "
               "legislative" if source_est_xlsx
               else "Bureaux de vote et circonscription (republie par un tiers)"),
    "source_url": "https://www.data.gouv.fr/datasets/circonscriptions-legislatives-table-de-correspondance/",
    "licence": "Licence Ouverte",
    "decoupage": "2010",
    "communes": table,
    "sans_circonscription": {c: (LIBELLES.get(c, "") if LIBELLES else "") for c in absentes},
}
brut = json.dumps(paquet, ensure_ascii=False, separators=(",", ":"))

# Relecture, sans reutiliser une variable d'au-dessus.
relu = json.loads(brut)
assert relu["v"] == 1
assert len(relu["communes"]) == len(par_commune)
assert all(isinstance(v, int) or (isinstance(v, list) and len(v) > 1)
           for v in relu["communes"].values()), "une forme intermediaire s'est glissee"
assert not (set(relu["communes"]) & set(relu["sans_circonscription"])), \
    "une commune est a la fois rattachee et declaree absente"

os.makedirs(os.path.dirname(OUT) or ".", exist_ok=True)
io.open(OUT, "w", encoding="utf-8").write(brut)

multi = [c for c, v in table.items() if isinstance(v, list)]
print("source                  : %s" % os.path.basename(SRC))
print("lignes lues             : %d" % n_lues)
print("communes rattachees     : %d" % len(par_commune))
if REF is not None:
    print("couverture              : %.2f %% des %d communes connues" % (100 * couverture, len(REF)))
    print("sans circonscription    : %d" % len(absentes))
    for c in absentes[:12]:
        print("   %s %s" % (c, (LIBELLES or {}).get(c, noms.get(c, ""))))
print("a cheval sur plusieurs  : %d (dont %s)"
      % (len(multi), ", ".join("%s : %d" % (noms.get(c, c), len(table[c]))
                               for c in sorted(multi, key=lambda x: -len(table[x]))[:3])))
if inconnues:
    print("codes non reconnus      : %d (communes fusionnees depuis 2017)" % sum(inconnues.values()))
if ecartees:
    print("ecartees volontairement : %s" % dict(ecartees))
print("poids                   : %.0f Ko" % (len(brut.encode()) / 1024))
