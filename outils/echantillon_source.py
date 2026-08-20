#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Decrit un fichier de donnees telecharge, quel que soit son format.

MEME RAISON QUE echantillon_scrutins.py : mon conteneur ne peut joindre ni
data.assemblee-nationale.fr ni data.gouv.fr (403 au tunnel). Le runner, lui, les
telecharge. Plutot que d'ecrire un lecteur sur des colonnes supposees, on fait
decrire le fichier par celui qui l'a sous la main.

Formats reconnus : CSV/TSV (delimiteur devine), XLSX (via openpyxl si present),
ZIP (listing), JSON. Pour tout autre : les premiers octets, en clair.

Usage :  python3 outils/echantillon_source.py data/circos_bureaux_de_vote.csv docs/schema_circos.md
"""
import sys, io, os, csv, json, zipfile, collections

SRC = sys.argv[1]
OUT = sys.argv[2]

def entete(s, titre):
    s.write("## %s\n\n" % titre)

def decrire_csv(chemin, s):
    brut = io.open(chemin, "rb").read(200000)
    # Le delimiteur se devine sur l'en-tete : celui qui la coupe en le plus de morceaux.
    tete = brut.split(b"\n")[0].decode("utf-8", "replace")
    delim = max([",", ";", "\t", "|"], key=lambda d: tete.count(d))
    s.write("- Delimiteur devine : `%s` (l'en-tete en contient %d)\n"
            % ({"\t": "TAB"}.get(delim, delim), tete.count(delim)))
    with io.open(chemin, encoding="utf-8", errors="replace", newline="") as f:
        lect = csv.reader(f, delimiter=delim)
        try:
            cols = next(lect)
        except StopIteration:
            s.write("\nFichier vide.\n"); return
        lignes = []
        vus = [collections.Counter() for _ in cols]
        total = 0
        for i, l in enumerate(lect):
            total += 1
            if i < 5:
                lignes.append(l)
            if i < 20000:
                for j, v in enumerate(l[:len(cols)]):
                    vus[j][v] += 1
        s.write("- Colonnes : **%d**\n- Lignes (hors en-tete) : **%d**\n\n" % (len(cols), total))
        s.write("| # | colonne | valeurs distinctes (sur 20 000 max) | exemples |\n|---|---|---|---|\n")
        for j, c in enumerate(cols):
            ex = " · ".join(str(v)[:40].replace("|", "\\|")
                            for v, _ in vus[j].most_common(3) if v != "")
            s.write("| %d | `%s` | %d | %s |\n" % (j, c, len(vus[j]), ex))
        s.write("\n### Cinq premieres lignes\n\n```\n")
        s.write(delim.join(cols) + "\n")
        for l in lignes:
            s.write(delim.join(l) + "\n")
        s.write("```\n")

def decrire_xlsx(chemin, s):
    try:
        import openpyxl
    except ImportError:
        s.write("openpyxl absent : impossible de lire le XLSX. "
                "Ajouter `pip install openpyxl` a la chaine.\n")
        # On decrit au moins l'archive : un XLSX est un ZIP.
        decrire_zip(chemin, s)
        return
    cl = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    s.write("- Feuilles : %s\n\n" % ", ".join("`%s`" % n for n in cl.sheetnames))
    for nom in cl.sheetnames[:3]:
        f = cl[nom]
        s.write("### Feuille `%s`\n\n```\n" % nom)
        for i, ligne in enumerate(f.iter_rows(values_only=True)):
            if i >= 8:
                break
            s.write(" | ".join("" if v is None else str(v)[:32] for v in ligne[:14]) + "\n")
        s.write("```\n\n")

def decrire_zip(chemin, s):
    with zipfile.ZipFile(chemin) as z:
        noms = z.namelist()
        s.write("- Entrees : **%d**\n\n```\n" % len(noms))
        for n in noms[:40]:
            s.write(n + "\n")
        if len(noms) > 40:
            s.write("... (%d de plus)\n" % (len(noms) - 40))
        s.write("```\n")

def decrire_json(chemin, s):
    d = json.load(io.open(chemin, encoding="utf-8"))
    s.write("```json\n" + json.dumps(d, ensure_ascii=False, indent=2)[:4000] + "\n```\n")

os.makedirs(os.path.dirname(OUT) or ".", exist_ok=True)
with io.open(OUT, "w", encoding="utf-8") as s:
    s.write("# Description de `%s`\n\n" % os.path.basename(SRC))
    s.write("Engendre par `outils/echantillon_source.py`. **Ne pas modifier a la main.**\n\n")
    if not os.path.exists(SRC):
        s.write("Fichier absent : la collecte ne l'a pas telecharge.\n")
        print("source absente : %s" % SRC)
        sys.exit(0)
    s.write("- Taille : **%.2f Mo**\n" % (os.path.getsize(SRC) / 1048576))
    bas = SRC.lower()
    tete = io.open(SRC, "rb").read(4)
    entete(s, "Contenu")
    if bas.endswith((".csv", ".tsv", ".txt")):
        decrire_csv(SRC, s)
    elif bas.endswith((".xlsx", ".xlsm")):
        decrire_xlsx(SRC, s)
    elif bas.endswith(".json"):
        decrire_json(SRC, s)
    elif tete[:2] == b"PK":
        decrire_zip(SRC, s)
    else:
        s.write("Format non reconnu. Premiers octets :\n\n```\n%s\n```\n"
                % io.open(SRC, "rb").read(600).decode("utf-8", "replace"))
print("description ecrite : %s" % OUT)
